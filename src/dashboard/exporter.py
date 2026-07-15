"""
src/dashboard/exporter.py
------------------------------------------------------------
Export filtered messages to CSV (StringIO) or Excel (BytesIO).

Cross-group forward detection (two-pass, priority order):

  Pass 1 -- fwd_from ground truth:
    If message.forward_from_id is set (Telethon fwd_from.from_id),
    the same original channel ID appearing in 2+ destination groups
    is a confirmed cross-group forward. This is the most reliable signal.

  Pass 2 -- normalised text fallback:
    For privacy-protected forwards where forward_from_id is NULL,
    text is normalised (strip + lower + collapse whitespace) before
    building the group-set map. This catches copy-pasted or near-identical
    messages that Telethon cannot attribute to an original source.

Excel highlight priority (highest wins):
  AMBER/GOLD  -- cross-group forward (either pass)
  GREEN       -- keyword hit only
  DARK AMBER  -- both
------------------------------------------------------------
"""

import csv
import io
import logging
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

_COLUMNS = [
    "id", "message_id", "group_name", "sender_name", "sender_phone",
    "text", "language", "is_forwarded", "matched_keyword", "fuzzy_score",
    "is_matched", "cross_group_forward", "timestamp",
]

_HEADER_LABELS = {
    "id":                  "Row",
    "message_id":          "Message ID",
    "group_name":          "Group",
    "sender_name":         "Sender",
    "sender_phone":        "Phone",
    "text":                "Message Text",
    "language":            "Language",
    "is_forwarded":        "Forwarded",
    "matched_keyword":     "Matched Keyword",
    "fuzzy_score":         "Fuzzy Score",
    "is_matched":          "Keyword Hit",
    "cross_group_forward": "Cross-Group Forward",
    "timestamp":           "Timestamp (UTC)",
}


# ---------------------------------------------------------------------------
# Cross-group forward detection
# ---------------------------------------------------------------------------

import re as _re


def _normalise(text: str) -> str:
    """
    Normalise message text before cross-group comparison.
    Handles: trailing spaces, different quote chars, collapsed whitespace,
    lowercase -- so 'Buy BTC now!' and 'buy btc now!' match.
    """
    if not text:
        return ""
    t = text.strip().lower()
    t = _re.sub(r"[\u2018\u2019\u201c\u201d]", "'", t)   # smart quotes -> straight
    t = _re.sub(r"\s+", " ", t)                           # collapse whitespace
    return t


def _find_cross_group_forwards(rows: list) -> set:
    """
    Return a set of row `hash` values that are confirmed cross-group forwards.

    Pass 1 -- fwd_from ground truth (forward_from_id):
      Same original channel ID forwarded into 2+ destination groups.
      Most reliable -- uses Telethon's fwd_from.from_id directly.

    Pass 2 -- normalised text fallback:
      Used when forward_from_id is NULL (privacy-protected forward, or
      message was copy-pasted rather than formally forwarded).
      text.strip().lower() + whitespace collapse before grouping.
    """
    from collections import defaultdict

    # Pass 1: fwd_from channel ID -> {dest group_ids}
    fwd_id_to_groups: dict = defaultdict(set)
    fwd_id_to_hashes: dict = defaultdict(list)
    for row in rows:
        if not row.get("is_forwarded"):
            continue
        src = row.get("forward_from_id")
        if not src:
            continue
        gid = row.get("group_id") or "unknown"
        fwd_id_to_groups[src].add(gid)
        fwd_id_to_hashes[src].append(row["hash"])

    cross_hashes: set = set()
    for src, groups in fwd_id_to_groups.items():
        if len(groups) >= 2:
            cross_hashes.update(fwd_id_to_hashes[src])

    # Pass 2: normalised text fallback (only rows not already caught by Pass 1)
    text_to_groups: dict = defaultdict(set)
    text_to_hashes: dict = defaultdict(list)
    for row in rows:
        if row["hash"] in cross_hashes:    # already flagged
            continue
        if not row.get("is_forwarded"):
            continue
        if row.get("forward_from_id"):      # has fwd_from but only 1 group -- skip
            continue
        norm = _normalise(row.get("text") or "")
        if len(norm) < 10:                  # too short to be meaningful
            continue
        gid = row.get("group_id") or "unknown"
        text_to_groups[norm].add(gid)
        text_to_hashes[norm].append(row["hash"])

    for norm, groups in text_to_groups.items():
        if len(groups) >= 2:
            cross_hashes.update(text_to_hashes[norm])

    return cross_hashes


def _annotate_rows(rows: list) -> list:
    """Add 'cross_group_forward' key (0/1) to each row dict."""
    # Ensure every row has a 'hash' key (backfilled rows may lack it)
    for row in rows:
        if "hash" not in row:
            row["hash"] = str(row.get("id", id(row)))

    cross_hashes = _find_cross_group_forwards(rows)
    for row in rows:
        row["cross_group_forward"] = 1 if row["hash"] in cross_hashes else 0
    return rows


# ---------------------------------------------------------------------------
# DB fetch
# ---------------------------------------------------------------------------

def _fetch_all(db_handler, filters: dict) -> list:
    result = db_handler.get_messages(
        keyword=filters.get("keyword"),
        group_id=filters.get("group_id"),
        datetime_from=filters.get("datetime_from") or filters.get("date_from"),
        datetime_to=filters.get("datetime_to") or filters.get("date_to"),
        matched_only=filters.get("matched_only", False),
        fetched_by=filters.get("fetched_by"),
        page=1,
        page_size=100_000,
    )
    return _annotate_rows(result["messages"])



# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def export_csv(db_handler, filters: dict) -> io.StringIO:
    rows = _fetch_all(db_handler, filters)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_COLUMNS, extrasaction="ignore")
    # header row with human labels
    writer.writerow({col: _HEADER_LABELS[col] for col in _COLUMNS})
    for row in rows:
        out = {col: row.get(col, "") for col in _COLUMNS}
        out["is_forwarded"]        = "Yes" if row.get("is_forwarded") else "No"
        out["is_matched"]          = "Yes" if row.get("is_matched")   else "No"
        out["cross_group_forward"] = "Yes" if row.get("cross_group_forward") else "No"
        writer.writerow(out)
    buf.seek(0)
    logger.info("CSV export: %d rows (%d cross-group forwards)",
                len(rows), sum(1 for r in rows if r.get("cross_group_forward")))
    return buf


# ---------------------------------------------------------------------------
# Excel styles
# ---------------------------------------------------------------------------

_HEADER_FILL      = PatternFill("solid", fgColor="3B1F6E")   # deep purple
_MATCH_FILL       = PatternFill("solid", fgColor="1A3A1A")   # dark green  -- keyword hit
_CROSS_GRP_FILL   = PatternFill("solid", fgColor="4A3000")   # amber/gold  -- cross-group forward
_BOTH_FILL        = PatternFill("solid", fgColor="3A2A00")   # darker amber -- hit AND cross-group

_HEADER_FONT  = Font(name="Calibri", bold=True,  color="FFFFFF", size=11)
_NORMAL_FONT  = Font(name="Calibri", bold=False, color="FFFFFF", size=10)
_CROSS_FONT   = Font(name="Calibri", bold=True,  color="FFD700", size=10)  # gold text

_THIN_BORDER = Border(
    left=Side(style="thin",  color="444444"),
    right=Side(style="thin", color="444444"),
    top=Side(style="thin",   color="444444"),
    bottom=Side(style="thin",color="444444"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

_COLUMN_WIDTHS = {
    "id": 6, "message_id": 14, "group_name": 22, "sender_name": 18,
    "sender_phone": 16, "text": 55, "language": 10, "is_forwarded": 12,
    "matched_keyword": 20, "fuzzy_score": 12, "is_matched": 12,
    "cross_group_forward": 20, "timestamp": 24,
}


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel(db_handler, filters: dict) -> io.BytesIO:
    rows = _fetch_all(db_handler, filters)

    cross_count = sum(1 for r in rows if r.get("cross_group_forward"))
    hit_count   = sum(1 for r in rows if r.get("is_matched"))

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ sheet 1
    ws = wb.active
    ws.title = "Messages"

    # Header
    for col_idx, col in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_HEADER_LABELS[col])
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _THIN_BORDER
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        is_hit   = bool(row.get("is_matched"))
        is_cross = bool(row.get("cross_group_forward"))

        # Determine fill + font priority
        # Cross-group forward wins over plain keyword hit
        if is_cross and is_hit:
            row_fill = _BOTH_FILL
            row_font = _CROSS_FONT
        elif is_cross:
            row_fill = _CROSS_GRP_FILL
            row_font = _CROSS_FONT
        elif is_hit:
            row_fill = _MATCH_FILL
            row_font = _NORMAL_FONT
        else:
            row_fill = None
            row_font = _NORMAL_FONT

        for col_idx, col in enumerate(_COLUMNS, start=1):
            val = row.get(col, "")
            if col in ("is_forwarded", "is_matched", "cross_group_forward"):
                val = "Yes" if val else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = row_font
            cell.border    = _THIN_BORDER
            cell.alignment = _WRAP if col == "text" else _CENTER
            if row_fill:
                cell.fill = row_fill

    # Column widths
    for col_idx, col in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COLUMN_WIDTHS.get(col, 15)

    # ------------------------------------------------------------------ sheet 2: Legend
    ws_legend = wb.create_sheet("Legend")
    ws_legend.column_dimensions["A"].width = 28
    ws_legend.column_dimensions["B"].width = 55

    legend_rows = [
        ("Colour", "Meaning"),
        ("Gold / Amber row",    "Same message forwarded to 2+ groups -- likely viral/spam content"),
        ("Green row",           "Message matched a tracked keyword (fuzzy score >= 85)"),
        ("Gold + Green (dark)", "Both: keyword hit AND cross-group forward -- highest priority"),
        ("No fill",             "Normal message -- no keyword match, not a cross-group forward"),
    ]
    for r_idx, (label, meaning) in enumerate(legend_rows, start=1):
        ws_legend.cell(row=r_idx, column=1, value=label).font = Font(bold=(r_idx == 1), size=10)
        ws_legend.cell(row=r_idx, column=2, value=meaning).font = Font(size=10)

    # ------------------------------------------------------------------ sheet 3: Summary
    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ("Export generated",           datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Total rows",                 len(rows)),
        ("Keyword hits",               hit_count),
        ("Cross-group forwards",       cross_count),
        ("Keyword hit + cross-group",  sum(1 for r in rows if r.get("is_matched") and r.get("cross_group_forward"))),
    ]
    for r_idx, (label, value) in enumerate(summary_data, start=1):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=True, size=10)
        ws2.cell(row=r_idx, column=2, value=value).font = Font(size=10)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(
        "Excel export: %d rows | %d keyword hits | %d cross-group forwards",
        len(rows), hit_count, cross_count,
    )
    return buf
